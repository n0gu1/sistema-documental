from datetime import datetime, time, timedelta
from io import BytesIO

from django.db.models import Prefetch
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from rest_framework import status
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response
from rest_framework.views import APIView

from .management_views import require_permission
from .auth_utils import record_auth_event
from .models import (
    ArchivoDocumento,
    DetalleSolicitudRevision,
    Documento,
    ProgramacionReporte,
    ReporteGenerado,
    SolicitudRevision,
)
from .permissions import IsAuthenticatedAndPasswordCurrent
from .reader_access import filter_accessible_documents


SCOPES = {'executive', 'editor', 'reviewer'}
FORMATS = {'PDF', 'XLSX'}
FREQUENCIES = {'daily', 'weekly', 'monthly'}
REPORT_DOCUMENT_PERMISSIONS = {
    'executive': 'reportes.generar',
    'editor': 'documentos.consultar',
    'reviewer': 'revisiones.consultar',
}


def parse_report_datetime(value, field_name, end=False):
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(str(value))
    except ValueError as error:
        raise ValidationError({field_name: 'Use una fecha ISO valida.'}) from error
    if parsed.tzinfo is None:
        parsed = timezone.make_aware(datetime.combine(parsed.date(), time.max if end else time.min))
    return parsed


def clean_filters(params):
    filters = {}
    for key in ('date_from', 'date_to', 'area_id', 'type_id', 'status_code', 'responsible_id'):
        value = params.get(key)
        if value not in (None, ''):
            filters[key] = str(value)
    parse_report_datetime(filters.get('date_from'), 'date_from')
    parse_report_datetime(filters.get('date_to'), 'date_to', end=True)
    return filters


def require_report_access(request, scope, generate=False):
    permission = {
        'executive': 'reportes.generar' if generate else 'reportes.generar',
        'editor': 'documentos.consultar',
        'reviewer': 'revisiones.consultar',
    }.get(scope)
    if not permission:
        raise ValidationError({'scope': 'El alcance del reporte no es valido.'})
    require_permission(request, permission)


def record_report_event(request, action_code, resource_id=None, details=None):
    record_auth_event(
        action_code=action_code,
        resource_code='REPORTE',
        organization_id=request.user.organizacion_id,
        user_id=request.user.id,
        session_id=getattr(request.auth, 'id', None),
        resource_id=resource_id,
        request=request,
        successful=True,
        result='Operacion de reporte correcta',
        details=details,
    )


def current_report_version(document):
    versions = getattr(document, 'report_versions', [])
    return versions[0] if versions else None


def document_report_rows(request, scope, filters):
    queryset = Documento.objects.filter(
        organizacion_id=request.user.organizacion_id,
        eliminado_en__isnull=True,
    ).select_related('area', 'tipo_documento', 'creado_por').prefetch_related(
        Prefetch(
            'archivos',
            queryset=ArchivoDocumento.objects.select_related('estado_version').filter(es_vigente=True),
            to_attr='report_versions',
        ),
    )
    if scope == 'editor':
        queryset = queryset.filter(creado_por_id=request.user.id)
    date_from = parse_report_datetime(filters.get('date_from'), 'date_from')
    date_to = parse_report_datetime(filters.get('date_to'), 'date_to', end=True)
    if date_from:
        queryset = queryset.filter(actualizado_en__gte=date_from)
    if date_to:
        queryset = queryset.filter(actualizado_en__lte=date_to)
    if filters.get('area_id'):
        queryset = queryset.filter(area_id=filters['area_id'])
    if filters.get('type_id'):
        queryset = queryset.filter(tipo_documento_id=filters['type_id'])
    if filters.get('responsible_id'):
        queryset = queryset.filter(creado_por_id=filters['responsible_id'])

    queryset = filter_accessible_documents(
        request.user,
        queryset,
        REPORT_DOCUMENT_PERMISSIONS[scope],
    )
    rows = []
    for document in queryset.order_by('-actualizado_en', 'codigo'):
        version = current_report_version(document)
        status_code = version.estado_version.codigo if version else 'SIN_VERSION'
        if filters.get('status_code') and status_code != filters['status_code']:
            continue
        rows.append({
            'id': str(document.id),
            'code': document.codigo,
            'title': document.nombre,
            'area_id': str(document.area_id),
            'area': document.area.nombre,
            'type_id': str(document.tipo_documento_id),
            'type': document.tipo_documento.nombre,
            'responsible_id': str(document.creado_por_id),
            'responsible': f'{document.creado_por.nombres} {document.creado_por.apellidos}'.strip(),
            'status_code': status_code,
            'status': version.estado_version.nombre if version else 'Sin version',
            'version': f'{version.numero_mayor}.{version.numero_menor}' if version else None,
            'updated_at': document.actualizado_en,
        })
    return rows


def reviewer_report_rows(request, filters):
    queryset = SolicitudRevision.objects.filter(
        revisor_id=request.user.id,
        version_documento__documento__organizacion_id=request.user.organizacion_id,
    ).select_related(
        'version_documento__documento__area',
        'version_documento__documento__tipo_documento',
        'version_documento__documento__creado_por',
        'estado_revision',
    ).prefetch_related('detalle').order_by('-solicitada_en')
    date_from = parse_report_datetime(filters.get('date_from'), 'date_from')
    date_to = parse_report_datetime(filters.get('date_to'), 'date_to', end=True)
    if date_from:
        queryset = queryset.filter(solicitada_en__gte=date_from)
    if date_to:
        queryset = queryset.filter(solicitada_en__lte=date_to)
    rows = []
    now = timezone.now()
    for review in queryset:
        document = review.version_documento.documento
        if not filter_accessible_documents(
            request.user,
            [document],
            REPORT_DOCUMENT_PERMISSIONS['reviewer'],
        ):
            continue
        detail = getattr(review, 'detalle', None)
        status_code = review.estado_revision.codigo
        if filters.get('area_id') and str(document.area_id) != filters['area_id']:
            continue
        if filters.get('type_id') and str(document.tipo_documento_id) != filters['type_id']:
            continue
        if filters.get('status_code') and status_code != filters['status_code']:
            continue
        if filters.get('responsible_id') and str(document.creado_por_id) != filters['responsible_id']:
            continue
        rows.append({
            'id': str(review.id),
            'code': document.codigo,
            'title': document.nombre,
            'area_id': str(document.area_id),
            'area': document.area.nombre,
            'type_id': str(document.tipo_documento_id),
            'type': document.tipo_documento.nombre,
            'responsible_id': str(document.creado_por_id),
            'responsible': f'{document.creado_por.nombres} {document.creado_por.apellidos}'.strip(),
            'status_code': status_code,
            'status': review.estado_revision.nombre,
            'deadline': detail.fecha_limite if detail else None,
            'overdue': bool(detail and detail.fecha_limite and detail.fecha_limite < now and not review.resuelta_en),
            'created_at': review.solicitada_en,
        })
    return rows


def report_options(rows):
    def unique(key, label_key):
        values = {row[key]: row[label_key] for row in rows if row.get(key) and row.get(label_key)}
        return [{'id': key, 'name': value} for key, value in sorted(values.items(), key=lambda item: item[1])]

    statuses = {row['status_code']: row['status'] for row in rows if row.get('status_code')}
    return {
        'areas': unique('area_id', 'area'),
        'types': unique('type_id', 'type'),
        'responsibles': unique('responsible_id', 'responsible'),
        'statuses': [{'id': key, 'name': value} for key, value in sorted(statuses.items())],
    }


def summarize_report(rows, scope):
    status_counts = {}
    area_counts = {}
    type_counts = {}
    responsible_counts = {}
    overdue = 0
    for row in rows:
        status_counts[row['status_code']] = status_counts.get(row['status_code'], 0) + 1
        area_counts[row['area']] = area_counts.get(row['area'], 0) + 1
        type_counts[row['type']] = type_counts.get(row['type'], 0) + 1
        responsible_counts[row['responsible']] = responsible_counts.get(row['responsible'], 0) + 1
        overdue += int(row.get('overdue', False))
    completed_statuses = {'APROBADO', 'PUBLICADO', 'COMPLETADA', 'APROBADA'}
    return {
        'total': len(rows),
        'published': status_counts.get('PUBLICADO', 0),
        'in_review': status_counts.get('EN_REVISION', 0) + status_counts.get('PENDIENTE', 0),
        'completed': sum(value for key, value in status_counts.items() if key in completed_statuses),
        'overdue': overdue,
        'by_status': [{'code': key, 'name': key.replace('_', ' ').title(), 'count': value} for key, value in sorted(status_counts.items())],
        'by_area': [{'name': key, 'count': value} for key, value in sorted(area_counts.items())],
        'by_type': [{'name': key, 'count': value} for key, value in sorted(type_counts.items())],
        'by_responsible': [{'name': key, 'count': value} for key, value in sorted(responsible_counts.items())],
        'scope': scope,
    }


def build_report_data(request, scope, filters):
    rows = reviewer_report_rows(request, filters) if scope == 'reviewer' else document_report_rows(request, scope, filters)
    return {
        'scope': scope,
        'filters': filters,
        'summary': summarize_report(rows, scope),
        'options': report_options(rows),
        'rows': rows,
    }


def serialize_report(report):
    return {
        'id': str(report.id),
        'name': report.nombre,
        'scope': report.alcance,
        'format': report.formato,
        'filters': report.filtros,
        'rows': report.filas,
        'created_at': report.creado_en,
        'download_url': f'/api/reports/{report.id}/download/',
    }


def report_history(request, scope):
    return [serialize_report(report) for report in ReporteGenerado.objects.filter(
        organizacion_id=request.user.organizacion_id,
        alcance=scope,
    )[:20]]


def cell_value(value):
    if isinstance(value, (datetime,)):
        return timezone.localtime(value).strftime('%Y-%m-%d %H:%M')
    return '' if value is None else value


def report_headers(scope):
    if scope == 'reviewer':
        return ['Código', 'Documento', 'Área', 'Tipo', 'Responsable', 'Estado', 'Vencida', 'Fecha']
    return ['Código', 'Documento', 'Área', 'Tipo', 'Responsable', 'Estado', 'Versión', 'Actualización']


def report_row_values(row, scope):
    if scope == 'reviewer':
        return [row['code'], row['title'], row['area'], row['type'], row['responsible'], row['status'], 'Sí' if row['overdue'] else 'No', cell_value(row['created_at'])]
    return [row['code'], row['title'], row['area'], row['type'], row['responsible'], row['status'], row['version'] or '', cell_value(row['updated_at'])]


def build_xlsx(data):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Resumen'
    sheet.append(['Reporte', 'Alcance', 'Generado'])
    sheet.append(['Reporte documental', data['scope'], timezone.localtime().strftime('%Y-%m-%d %H:%M')])
    sheet.append([])
    sheet.append(['Indicador', 'Valor'])
    for key, value in data['summary'].items():
        if isinstance(value, (int, float)):
            sheet.append([key, value])
    for cell in sheet[1] + sheet[4]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1F4E78')
    detail = workbook.create_sheet('Detalle')
    detail.append(report_headers(data['scope']))
    for row in data['rows']:
        detail.append(report_row_values(row, data['scope']))
    for sheet_item in workbook.worksheets:
        for column in sheet_item.columns:
            width = min(max(len(str(cell.value or '')) for cell in column) + 2, 42)
            sheet_item.column_dimensions[column[0].column_letter].width = width
        sheet_item.freeze_panes = 'A2'
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_pdf(data):
    output = BytesIO()
    document = SimpleDocTemplate(output, pagesize=landscape(letter), rightMargin=0.35 * inch, leftMargin=0.35 * inch, topMargin=0.35 * inch, bottomMargin=0.35 * inch)
    styles = getSampleStyleSheet()
    story = [Paragraph('Reporte documental', styles['Title']), Paragraph(f"Alcance: {data['scope']} | Registros: {len(data['rows'])}", styles['Normal']), Spacer(1, 12)]
    summary = [['Indicador', 'Valor']] + [[key, value] for key, value in data['summary'].items() if isinstance(value, (int, float))]
    summary_table = Table(summary, colWidths=[2.2 * inch, 1 * inch])
    summary_table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white), ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D9E2F3')), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold')]))
    story.extend([summary_table, Spacer(1, 14)])
    detail = [report_headers(data['scope'])] + [[str(value) for value in report_row_values(row, data['scope'])] for row in data['rows'][:1000]]
    detail_table = Table(detail, repeatRows=1)
    detail_table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white), ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D9E2F3')), ('FONTSIZE', (0, 0), (-1, -1), 7), ('VALIGN', (0, 0), (-1, -1), 'TOP')]))
    story.append(detail_table)
    document.build(story)
    return output.getvalue()


def report_response(data, report_format):
    content = build_pdf(data) if report_format == 'PDF' else build_xlsx(data)
    content_type = 'application/pdf' if report_format == 'PDF' else 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    extension = report_format.lower()
    response = HttpResponse(content, content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="reporte-documental.{extension}"'
    return response


class ReportListView(APIView):
    permission_classes = [IsAuthenticatedAndPasswordCurrent]

    def get(self, request):
        scope = request.query_params.get('scope', 'executive')
        require_report_access(request, scope)
        filters = clean_filters(request.query_params)
        data = build_report_data(request, scope, filters)
        data['history'] = report_history(request, scope)
        record_report_event(request, 'REPORTE_CONSULTADO', details={'scope': scope, 'rows': len(data['rows'])})
        return Response(data)


class ReportGenerateView(APIView):
    permission_classes = [IsAuthenticatedAndPasswordCurrent]

    def post(self, request):
        scope = request.data.get('scope', 'executive')
        require_report_access(request, scope, generate=True)
        report_format = str(request.data.get('format', 'PDF')).upper()
        if report_format not in FORMATS:
            raise ValidationError({'format': 'El formato debe ser PDF o XLSX.'})
        filters = clean_filters(request.data.get('filters', {}))
        data = build_report_data(request, scope, filters)
        report = ReporteGenerado.objects.create(
            organizacion_id=request.user.organizacion_id,
            generado_por_id=request.user.id,
            alcance=scope,
            formato=report_format,
            nombre=f'Reporte {scope} - {timezone.localtime().strftime("%Y-%m-%d %H:%M")}',
            filtros=filters,
            filas=len(data['rows']),
        )
        record_report_event(
            request,
            'REPORTE_GENERADO',
            resource_id=report.id,
            details={'scope': scope, 'format': report_format, 'rows': len(data['rows'])},
        )
        return Response({'report': serialize_report(report)}, status=status.HTTP_201_CREATED)


class ReportDownloadView(APIView):
    permission_classes = [IsAuthenticatedAndPasswordCurrent]

    def get(self, request, report_id):
        report = ReporteGenerado.objects.filter(pk=report_id, organizacion_id=request.user.organizacion_id).first()
        if not report:
            return Response({'detail': 'Reporte no encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        require_report_access(request, report.alcance)
        response = report_response(build_report_data(request, report.alcance, report.filtros), report.formato)
        record_report_event(request, 'REPORTE_DESCARGADO', resource_id=report.id, details={'format': report.formato})
        return response


def serialize_schedule(schedule):
    return {
        'id': str(schedule.id),
        'name': schedule.nombre,
        'scope': schedule.alcance,
        'format': schedule.formato,
        'frequency': schedule.frecuencia,
        'filters': schedule.filtros,
        'next_run_at': schedule.proxima_ejecucion_en,
        'active': schedule.activa,
    }


class ReportScheduleListView(APIView):
    permission_classes = [IsAuthenticatedAndPasswordCurrent]

    def get(self, request):
        scope = request.query_params.get('scope', 'executive')
        require_report_access(request, scope)
        schedules = ProgramacionReporte.objects.filter(organizacion_id=request.user.organizacion_id, alcance=scope, activa=True)
        response = Response({'schedules': [serialize_schedule(item) for item in schedules]})
        record_report_event(request, 'REPORTE_CONSULTADO', details={'scope': scope, 'scheduled': True})
        return response

    def post(self, request):
        scope = request.data.get('scope', 'executive')
        require_report_access(request, scope, generate=True)
        frequency = request.data.get('frequency', 'monthly')
        report_format = str(request.data.get('format', 'PDF')).upper()
        if frequency not in FREQUENCIES:
            raise ValidationError({'frequency': 'La frecuencia debe ser daily, weekly o monthly.'})
        if report_format not in FORMATS:
            raise ValidationError({'format': 'El formato debe ser PDF o XLSX.'})
        filters = clean_filters(request.data.get('filters', {}))
        next_run = parse_report_datetime(request.data.get('next_run_at'), 'next_run_at') or timezone.now() + timedelta(days=1)
        schedule = ProgramacionReporte.objects.create(
            organizacion_id=request.user.organizacion_id,
            creado_por_id=request.user.id,
            nombre=request.data.get('name') or f'Reporte {scope} programado',
            alcance=scope,
            formato=report_format,
            frecuencia=frequency,
            filtros=filters,
            proxima_ejecucion_en=next_run,
        )
        record_report_event(
            request,
            'REPORTE_PROGRAMADO',
            resource_id=schedule.id,
            details={'scope': scope, 'format': report_format, 'frequency': frequency},
        )
        return Response({'schedule': serialize_schedule(schedule)}, status=status.HTTP_201_CREATED)


class ReportScheduleDetailView(APIView):
    permission_classes = [IsAuthenticatedAndPasswordCurrent]

    def patch(self, request, schedule_id):
        schedule = ProgramacionReporte.objects.filter(pk=schedule_id, organizacion_id=request.user.organizacion_id).first()
        if not schedule:
            return Response({'detail': 'Programacion no encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        require_report_access(request, schedule.alcance, generate=True)
        if 'active' in request.data:
            schedule.activa = bool(request.data['active'])
        if 'next_run_at' in request.data:
            schedule.proxima_ejecucion_en = parse_report_datetime(request.data['next_run_at'], 'next_run_at')
        schedule.save(update_fields=['activa', 'proxima_ejecucion_en', 'actualizada_en'])
        record_report_event(request, 'REPORTE_PROGRAMACION_MODIFICADA', resource_id=schedule.id, details={'active': schedule.activa})
        return Response({'schedule': serialize_schedule(schedule)})

    def delete(self, request, schedule_id):
        schedule = ProgramacionReporte.objects.filter(pk=schedule_id, organizacion_id=request.user.organizacion_id).first()
        if not schedule:
            return Response({'detail': 'Programacion no encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        require_report_access(request, schedule.alcance, generate=True)
        schedule.activa = False
        schedule.save(update_fields=['activa', 'actualizada_en'])
        record_report_event(request, 'REPORTE_PROGRAMACION_CANCELADA', resource_id=schedule.id)
        return Response(status=status.HTTP_204_NO_CONTENT)
